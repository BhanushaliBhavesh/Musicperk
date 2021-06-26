import mysql.connector
from openpyxl import load_workbook, Workbook

mydb=mysql.connector.connect(host='localhost',user='root',password='admin')
mycur=mydb.cursor()
mycur.execute("CREATE DATABASE IF NOT EXISTS musicperk")
mydb=mysql.connector.connect(host='localhost',user='root',password='admin',database='musicperk')
cur=mydb.cursor()
query_1="CREATE TABLE IF NOT EXISTS hindalco(dat DATETIME, close FLOAT(10,2),high FLOAT(10,2),low FLOAT(10,2),open FLOAT(10,2),volume INTEGER(19),instrument CHAR(15))"
cur.execute(query_1)

#inserting data in database
wb = load_workbook('HINDALCO_1D.xlsx')
ws = wb.active
for row in ws.iter_rows(min_row=2,max_col=7):
    for cell in row:
        
        if cell.column == 1:
            dat= cell.value
        elif cell.column ==2:
            clo = cell.value
        elif cell.column ==3:
            high = cell.value
        elif cell.column ==4:
            low = cell.value
        elif cell.column ==5:
            open = cell.value
        elif cell.column ==6:
            volume = cell.value
        elif cell.column ==7:
            instrument = cell.value
    
    sql = "INSERT INTO hindalco(dat, close, high, low, open, volume, instrument) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    vlu = (dat,clo,high,low,open,volume,instrument)
    cur.execute(sql,vlu)

mydb.commit()
